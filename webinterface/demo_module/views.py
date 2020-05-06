"""
Contains view functions for the demo_module
URL paths that lead here are in demo_module/urls.py
"""

from django.shortcuts import render, render_to_response, redirect
from django.http import HttpResponse, HttpResponseRedirect
from django.views.generic import ListView
from django.template import loader
from datetime import datetime
from django.conf import settings
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
import locale


from comms.messagehandler.client import MqttClient
from comms.messagehandler import protocol
from .forms import TestForm, AccelerometerForm
from .models import Result, Status
from .models import Inbound_teststand_package, Test_stand_data, Test_stand_parameters, ND_TS

# Bokeh for charts
from bokeh.plotting import figure
from bokeh.embed import components
from bokeh.models import HoverTool, ResetTool, FreehandDrawTool

# Numpy for data
import numpy as np
from numpy import fft as fft

# CSV writer without writing to a file on disk
import csv
import io

import json

def gui_demo(request):
    """
    This view generates contents for a demo page, to show GUI possibilities.
    It is semi-functional, meant to be a cut & paste template for implementors.
    Janus, Feb-March 2020.
    """

    # Denne formular bliver vist på siden
    form = AccelerometerForm()

    # Denne data bliver visualiseret på siden
    N = 10000   # Num samples
    fs = 16000  # Sampling freq
    Ts = 1/fs   # Sampling time
    A = 5       # Amplitude
    f0 = 4000   # Hz

    t = np.linspace(0, N*Ts, N)
    n = np.arange(0, N)
    x = 1.2*A*np.sin(2*np.pi*f0/fs*n) + A*np.sin(2*np.pi*0.75*f0/fs*n) + 0.4*A*np.sin(2*np.pi*1.33*f0/fs*n)

    X = np.fft.fftshift( 20*np.log10( np.abs(fft.fft(x, N)) ) )
    f = np.fft.fftshift( fft.fftfreq(N, d=Ts) )


    plot = figure(title='FFT powerspektrum 😍',
                  sizing_mode='scale_width',
                  x_axis_label='f [Hz]', y_axis_label='|X(f)|^2 [dB]',
                  x_range = [0, fs/2],
                  plot_width=800, toolbar_location="below")

    plot.add_tools(HoverTool())
    plot.line(f, X, legend_label='Powerspektrum for x(t)', color='blue')

    # Her bliver figuren lavet til hhv. JavaScript og indhold til et HTML-div
    script, div = components(plot)

    # Her bliver siden kaldt og variablerne  script, div og form bliver leveret med...
    return render(request, 'demo_module/gui-demo.html', {'script': script, 'div': div, 'form': form})


# Show landing page for the demo module
def demo_main_page(request):
    return render(request, 'demo_module/home.html')


def demo_create_test_middle_link(request):
    # max_time_difference is the time a test, may maximum take to complete
    max_time_difference = 0.1
    mtds = max_time_difference * 60

    locale.setlocale(locale.LC_TIME, 'da_DK.utf8')

    # Check if there exist a ND_TS table
    # if not create one
    try:
        ND_TS.objects.all()[0]
        temp = ND_TS.objects.all()[0]
    except:
        temp = ND_TS()
        temp.TimeStamp = "01/01/2000-23:59:59"
        temp.ID = 0
        temp.save()

    # Last test time
    ts = temp.TimeStamp
    tsf = datetime.strptime(ts, "%d/%m/%Y-%H:%M:%S")

    # Current time
    timestamp = datetime.now()

    # Get the difference between them
    difference = timestamp - tsf

    # See if the last test is done (from guessed maximum test time)
    # If the last test is done, redirect to create test page
    # If not, give "error" message
    if difference.seconds > mtds:
        return redirect('demo_make_test')
    else:
        template = loader.get_template('demo_module/message_sent.html')
        outcome = "Test stand er i øjeblikket ikke ledig "
        context = {'outcome': outcome, }
        return HttpResponse(template.render(context, request))


# Show test creation form for the demo module
def demo_create_test(request):

    if request.method == 'POST':
        form = TestForm(request.POST)

        # Validate form
        if form.is_valid():
            # Do something with it, e.g. store it and send the MQTT message
            print("Form submitted!")

            template = loader.get_template('demo_module/message_sent.html')

            # ------- Temporary saving table ------#
            # save time sent
            locale.setlocale(locale.LC_TIME, 'da_DK.utf8')
            temp = ND_TS()
            timestamp = datetime.now()
            temp.TimeStamp = timestamp.strftime("%d/%m/%Y-%H:%M:%S")
            temp.ID = 0

            # save no delete field
            temp.NoDelete = form.cleaned_data.get("no_delete")
            temp.save()
            # ------- -------------------- ------#

            # Attempt to transmit MQTT-message based on validated form data
            if (transmit_mqtt(form.cleaned_data)):
                outcome = "Succes. Beskeden blev sendt."
                # Set "test stand" as not available
                temp = ND_TS.objects.all()[0]
                temp.Statusbool = False
                temp.save()
            else:
                outcome = "Fejl. Beskeden blev ikke."

            # Show result to user
            context = {'outcome': outcome, }
            return HttpResponse(template.render(context, request))

    # GET-request, possibly failed submission
    else:
        # instantiate a new form to pass into the template context
        form = TestForm()


    # Render the form template
    return render(request, 'demo_module/make_test.html', {'form': form})

# Callback on publishing - After handshakes
def on_publish_callback(client, userdata, mid):
    global sending
    sending = False

def transmit_mqtt(form_obj):
    """
    This function is not a view.
    This function transmits a validated message.
    Janus, March 2020
    """

    # Print to console for debug
    print(form_obj)

    # Create a message to send
    topic = form_obj['topic']
    # Payload
    m = protocol.Message()
    m.new()
    m.sentBy = form_obj['sender']
    m.msgType = form_obj['msg_type']
    m.statusCode = form_obj['status_code']

    # try-except on the json conversions
    # convert json -> python
    try:
        m.commandList = json.loads(form_obj['command_list_str'])
    except:
        print("Error converting commandlist -> insert empty")

    try:
        m.dataObj = json.loads(form_obj['data_obj_str'])
    except:
        print("Error converting dataObj -> insert empty")

    try:
        m.parameterObj = json.loads(form_obj['parameter_obj_str'])
    except:
        print("Error converting parameterObj -> insert empty")

    # Done inserting data
    m.pack()
    send_me = protocol.ProtocolSchema.write_jsonstr(m.payload)

    # debug output to console
    print(send_me)

    # Send it

    # The donothing callback function
    def donothing(client, userdata, message):
        pass

    # Callback on publishing - After handshakes
    def on_publish_callback(client, userdata, mid):
        global sending
        sending = False

    # Create client
    publisher = MqttClient("DemoModuleMessageSender", donothing, on_publish_callback)

    # Send and disconnect
    rc = publisher.publish(topic, send_me)

    publisher.loop_start()
    global sending
    sending = True
    # Wait for the handshaking to end
    while sending:
        pass
    publisher.loop_stop()

    publisher.disconnect()

    return rc


# show running test page will include webcam feed
def demo_running_test(request):
    #Not done
    return render(request, 'demo_module/running_test.html')



def show_data(request, test_id):
    """
    Show specific test data page, specific datapoints from a test from saved_data
    Janus, April 2020
    """

    # Fetch the specific test
    my_test = Inbound_teststand_package.objects.get(id=test_id)

    # Fetch the parameters
    my_params = Inbound_teststand_package.objects.get(id=test_id).parameters.all()

    # Fetch the data test and extract x-y
    my_dataset = Inbound_teststand_package.objects.get(id=test_id).data.all()
    x = my_dataset[0].Data_points
    y = my_dataset[1].Data_points

    # Make a time domain plot
    plot_time = figure(title='x - y(x) plot 🧐',
                  sizing_mode='scale_width',
                  x_axis_label='x [Antal flødeboller]', y_axis_label='|Ekstra vægt|^2 [kg]',
                  plot_width=800, toolbar_location="below")

    plot_time.add_tools(HoverTool())
    plot_time.line(x, y, legend_label='Ekstra vægt (kg)', color='blue')

    # Her bliver figuren lavet til hhv. JavaScript og indhold til et HTML-div
    script_time, div_time = components(plot_time)

    # Make an FFT plot
    fs = 100    # Assumed, this must be changed
    Ts = 1 / fs
    N = 1024 # len(y), eller Zero-padding
    Y = np.fft.fftshift(20 * np.log10(np.abs(fft.fft(y, N))))
    f = np.fft.fftshift(fft.fftfreq(N, d=Ts))
    plot_fft = figure(title='Effektspektrum |Y(f)|^2. Antaget fs = 100 Hz, N = 1024 🤯',
                       sizing_mode='scale_width',
                       x_axis_label='f [Hz]', y_axis_label='|Y(f)|^2 [dB]',
                       x_range=[0, fs / 2],
                       plot_width=800, toolbar_location="below")

    plot_fft.add_tools(HoverTool())
    plot_fft.line(f, Y, legend_label='Effektspektrum for y', color='blue')

    # Her bliver figuren lavet til hhv. JavaScript og indhold til et HTML-div
    script_fft, div_fft = components(plot_fft)

    template_name = 'demo_module/show_data.html'
    return render(request, template_name, {'script_time': script_time,
                                           'div_time': div_time,
                                           'script_fft': script_fft,
                                           'div_fft': div_fft,
                                           'test': my_test,
                                           'params': my_params,
                                           'data': my_dataset
                                           })

#All saved tests
class ResultListView(ListView):
    model = Inbound_teststand_package
    queryset = Inbound_teststand_package.objects.all().order_by('-Timestamp')

    def get_context_data(self, **kwargs):
        """
        This method overwrites the class-based in order to add more context.
        Janus, April 2020
        """
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        # Add in our extra
        context['dyn_url'] = self.request.build_absolute_uri('')
        context['dyn_csv_url'] = self.request.build_absolute_uri('csv')
        context['dyn_excel_url'] = self.request.build_absolute_uri('excel')
        return context


#Info page about the demo module
class StatusListView(ListView):
    model = Status
    queryset = Status.objects.all()


# busy page, test already running
def demo_busy(request):
    return render(request, 'demo_module/busy.html')


# Streaming page demo
def demo_stream(request):
    """
    This view must fetch the relevant streaming address and
    insert it into the tempate
    """

    # Get the DK cam settings from the Django settings file
    cam1 = settings.CAMS['cam1']
    cam1.update({
        'activate': request.build_absolute_uri(cam1['api_activate']),
        'url' : request.build_absolute_uri(cam1['api_url'])
    })

    # Get the Japan cam settings from the Django settings file
    cam2 = settings.CAMS['cam2']
    cam2.update({
        'activate': request.build_absolute_uri(cam2['api_activate']),
        'url' : request.build_absolute_uri(cam2['api_url'])
    })

    context = {'cam1': cam1, 'cam2': cam2}
    return render(request, 'demo_module/streaming-demo.html', context)

def make_csv_from_db(request, test_id):
    """
    This view function extracts a test result (pk=test_id) from the database and
    formats it into a structure that can be written to a CSV.
    It then writes a CSV and returns it as a Http response.
    Janus, April 2020
    """

    # Get the test result
    my_test = Inbound_teststand_package.objects.get(id=test_id)

    # Make a master data header and values
    masterdata_header = ['Testdato', 'Oprettet af', 'Kommandoliste']
    masterdata_values = [my_test.Timestamp, my_test.Sent_by, my_test.command_list]

    # Get parameters from the test
    my_params = Inbound_teststand_package.objects.get(id=test_id).parameters.all()
    csv_param_names = [p.Parameter_name for p in my_params]
    csv_param_values = [p.Parameter_value for p in my_params]

    # Get data points from the test
    my_dataset = Inbound_teststand_package.objects.get(id=test_id).data.all()

    # Header / column names for the CSV
    # ['x', 'y', 'z'] -> [('x','y','z')]
    csv_columns = [d.Data_name for d in my_dataset]
    #csv_columns = zip(*column_names)

    # All the data [[x1, x2, x3, ...], [y1, y2, y3, ...], [z1, z2, z3, ...], etc...]
    data_points = [d.Data_points for d in my_dataset]

    # All the data joined together per row [(x1,y1,z1), (x2,y2,z2), ...]
    csv_data = zip(*data_points)

    # We will write the CSV to this buffer instead of a file on disk
    buffer = io.StringIO()

    # Make a writer object, and put the values into the buffer stream
    csv_writer = csv.writer(buffer, dialect='excel')
    csv_writer.writerow(masterdata_header)
    csv_writer.writerow(masterdata_values)
    csv_writer.writerow(csv_param_names)
    csv_writer.writerow(csv_param_values)
    csv_writer.writerow(csv_columns)
    csv_writer.writerows(csv_data)

    # Rewind buffer
    buffer.seek(0)

    # Make the response from buffer and set MIME -> send to user
    response = HttpResponse(buffer, content_type='text/csv')
    return response

def make_excel_from_db(request, test_id):
    workbook = Workbook()

    # ------------------ data sheet ------------------#

    # Create sheet
    data_sheet = workbook.create_sheet("Data")
    #data_sheet = workbook.active

    # Get the test result
    my_test = Inbound_teststand_package.objects.get(id=test_id)

    # Make a master data header and values
    master_data_header = [
        ["Test dato", "Oprettet af"],
        [my_test.Timestamp, my_test.Sent_by],
        [" "],
    ]

    # Fill out header rows in spread sheet
    for row in master_data_header:
        data_sheet.append(row)

    # Get data from database
    my_data = Inbound_teststand_package.objects.get(id=test_id).data.all()
    data_names = [d.Data_name for d in my_data]
    data_points = [d.Data_points for d in my_data]

    # Fill out data row header colums
    for column, text in enumerate(data_names, start=1):
        data_sheet.cell(column=column, row=4, value=text)

    # Fill out data point rows
    for column, row_entries in enumerate(data_points, start=1):
        for row, value in enumerate(row_entries, start=5):
            data_sheet.cell(column=column, row=row, value=value)

    #------------------ Chart sheet ------------------#

    # Create new sheet
    chart_sheet = workbook.create_sheet("chart")

    # Init chart
    chart = LineChart()

    # Get chart data from data sheet
    chart_data = Reference(worksheet=data_sheet,
                           min_row=4,
                           max_row=row,
                           min_col=1,
                           max_col=column)

    # Make chart titles
    chart.title = "Tids domæne"
    chart.y_axis.title = "Vægt i Kg"
    chart.x_axis.title = "Antal flødeboller"

    # Create chart
    chart.add_data(chart_data, titles_from_data=True)
    chart_sheet.add_chart(chart, "A1")

    # ------------------ parameter sheet ------------------#

    # Create new sheet
    param_sheet = workbook.create_sheet("Parameters")

    # Get parameters from database
    my_param = Inbound_teststand_package.objects.get(id=test_id).parameters.all()
    param_names = [p.Parameter_name for p in my_param]
    param_points = [p.Parameter_value for p in my_param]

    # Fill out parameter header (names) row
    for column, text in enumerate(param_names, start=1):
        param_sheet.cell(column=column, row=1, value=text)

    # Fill out parameter values row
    for column, text in enumerate(param_points, start=1):
        param_sheet.cell(column=column, row=2, value=text)

    # ------------------ End game ------------------#

    # Remove extra sheet (dunno why it is there, but it just is?)
    sheet = workbook.get_sheet_by_name('Sheet')
    workbook.remove_sheet(sheet)

    # Create the Http response, and set content type to be excel spreadsheet
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    # Save the excel document as the request response
    workbook.save(response)

    # Return the excel document
    return response

# def send_mqtt(request):
#
#     # Create a message to send
#     topic = "demo_module/inbound"
#
#     # Payload
#     m = protocol.Message()
#     m.new()
#     m.sentBy = "Team 2"
#     m.msgType = "data"
#     m.statusCode = "200"
#     m.dataObj = {
#         'x': [1,2,3,4,5,6,7,8,9,10],
#         'y': [1,4,9,16,25,36,49,64,81,100]
#     }
#
#     # Get ready to send
#     m.pack()
#     send_me = protocol.ProtocolSchema.write_jsonstr(m.payload)
#
#     # The donothing function
#     def donothing(client, userdata, message):
#         pass
#
#     # Create client
#     publisher = MqttClient("MessageSender", donothing)
#
#     # Send and disconnect
#     rc = publisher.publish(topic, send_me)
#     publisher.disconnect()
#
#     # Render a response
#     template = loader.get_template('demo_module/message_sent.html')
#
#     if rc:
#         outcome = "Success. Message was sent."
#     else:
#         outcome = "Failure. Message was not sent."
#
#     context = {'outcome': outcome,
#                'topic': topic,
#                'msg': send_me}
#
#     return HttpResponse(template.render(context, request))
